import json
from typing import Dict, Union, BinaryIO

import openpyxl
from models.pdf_link import PdfLink


def save_links_into_excel(xlsx_path: str, links: Dict[str, PdfLink]) -> None:
    data = {k: v.to_json() for k, v in links.items()}
    json_text = json.dumps(data, ensure_ascii=False)

    wb = openpyxl.load_workbook(xlsx_path)

    if "__PDF_LINKS__" in wb.sheetnames:
        ws = wb["__PDF_LINKS__"]
    else:
        ws = wb.create_sheet("__PDF_LINKS__")

    # Clear existing contents
    if ws.max_row > 1 or ws.max_column > 1 or ws["A1"].value is not None:
        ws.delete_rows(1, ws.max_row)

    chunk_size = 30000
    for i in range(0, len(json_text), chunk_size):
        ws.cell(row=1 + i // chunk_size, column=1, value=json_text[i:i + chunk_size])

    ws.sheet_state = "hidden"
    wb.save(xlsx_path)
    wb.close()


def try_load_embedded_links(xlsx_source: Union[str, BinaryIO]) -> Dict[str, PdfLink]:
    try:
        if hasattr(xlsx_source, "seek"):
            xlsx_source.seek(0)

        wb = openpyxl.load_workbook(xlsx_source, read_only=True, data_only=True)

        if "__PDF_LINKS__" not in wb.sheetnames:
            wb.close()
            return {}

        ws = wb["__PDF_LINKS__"]
        parts = [str(row[0]) for row in ws.iter_rows(values_only=True) if row and row[0]]

        wb.close()

        if not parts:
            return {}

        data = json.loads("".join(parts))
        return {k: PdfLink.from_json(v) for k, v in data.items()}

    except Exception:
        return {}